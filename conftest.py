import os
import pytest
import allure
import openpyxl
from pathlib import Path
from datetime import datetime
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.chrome.service import Service as chrome_service
from selenium.webdriver.firefox.service import Service as firefox_service
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.safari.options import Options as SafariOptions

pytest_plugins = [
    "pages_fixtures",
]


def pytest_addoption(parser):
    parser.addoption('--browser', action='store',
                     default='chrome', help="browser to run tests on")
    parser.addoption(
        "--headless", action="store_true", default=False, help="hide tests execution"
    )


@pytest.fixture(scope="session", autouse=True)
def add_folders():
    """
    Generate the media folder if not present
    """
    dir = os.getenv("MEDIA_FOLDER", "media")
    Path(dir).mkdir(parents=True, exist_ok=True)


@pytest.fixture(scope="session", autouse=True)
def driver(request):
    """
    Build the driver to be used in the tests and then quit it when test execution is over
    """
    browser = request.config.getoption('browser').lower()
    driver = None
    available_options = {
        "chrome": ChromeOptions,
        "firefox": FirefoxOptions,
        "safari": SafariOptions

    }
    # Select options based on browser
    options = available_options.get(browser)
    if options:
        options = options()
        if request.config.getoption('headless'):
            options.add_argument("--headless")
    else:
        raise ValueError("Invalid browser selected")

    if browser == "chrome":
        driver = webdriver.Chrome(service=chrome_service(
            ChromeDriverManager().install()), options=options)
    elif browser == "firefox":
        # Firefox needs a special profile to handle cookies properly
        options.set_preference('profile', os.getenv("FIREFOX_PROFILE"))
        driver = webdriver.Firefox(
            service=firefox_service(GeckoDriverManager().install()), options=options)
    elif browser == "safari":
        driver = webdriver.Safari()
    driver.set_window_size(1920, 1080)
    yield driver
    driver.quit()


@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    """
    Hook to check if a test has failed
    """
    outcome = yield
    rep = outcome.get_result()
    # set a report attribute for each phase of a call (setup, call, teardown)
    setattr(item, "rep_" + rep.when, rep)


@pytest.fixture(scope="function", autouse=True)
def test_failed_check(request):
    """
    Add a screenshot to allure report if a test execution fails
    """
    yield
    used_browser = request.config.getoption("browser")
    allure.dynamic.parameter("browser", used_browser.capitalize())
    if request.node.rep_setup.failed:
        print("setting up a test failed!", request.node.nodeid)
    # Add screenshot only if failed during execution
    elif request.node.rep_setup.passed:
        if request.node.rep_call.failed:
            driver = request.node.funcargs['driver']
            file_path = take_screenshot(
                driver, request.node.nodeid, used_browser)
            add_to_allure(file_path)
            print("executing test failed", request.node.nodeid)


def take_screenshot(driver, nodeid, browser):
    """
    Take a screenshot of the browser and store it in the media folder
    """
    file_name = f'{nodeid}_{browser}_{datetime.today().strftime("%Y-%m-%d_%H:%M")}.png'.replace("/",
                                                                                                "_").replace("::", "__")
    file_full_path = os.path.join(
        os.getenv("MEDIA_FOLDER", "media"), file_name)
    driver.save_screenshot(file_full_path)
    return file_full_path


def add_to_allure(img_path):
    """
    Add screenshot to allure report
    """
    with open(img_path, "rb") as image:
        data = image.read()
        allure.attach(body=data, name="Screenshot",
                      attachment_type=allure.attachment_type.PNG)


def pytest_generate_tests(metafunc):
    """
    Generate parametrize tests from excel file.
    Data format: List of Tuple[List[str],int]
    The first element of the tuple are the filters to apply
    The second element of the tuple are the expected number of products that satisfy the filter
    """
    if "filter_data" in metafunc.fixturenames:
        workbook = openpyxl.load_workbook('data/filter_data.xlsx')
        worksheet = workbook.active
        tests_data = []
        headers = {position: cell.value for position,
                   cell in enumerate(worksheet[1])}
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            cells = [cell for cell in row]
            tests_data.append(
                (
                    {headers[position]: str(cell.value).split(",") for position,
                     cell in enumerate(cells[:-1]) if cell.value},
                    cells[-1].value)
            )
        metafunc.parametrize("filter_data", tests_data)
